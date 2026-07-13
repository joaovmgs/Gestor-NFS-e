from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from tempfile import NamedTemporaryFile
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


NAMESPACE = {"n": "http://www.sped.fazenda.gov.br/nfse"}
MONEY_FORMAT = 'R$ #,##0.00;[Red]-R$ #,##0.00'


@dataclass(frozen=True)
class ReportRow:
    numero: str
    situacao: str
    finalidade_nfse: str
    tipo_debito: str
    tipo_credito: str
    valor_total: Decimal | None
    irrf_retido: Decimal | None
    csll_retida: Decimal | None
    issqn_retido: Decimal | None
    contribuicao_previdenciaria_retida: Decimal | None
    valor_ibs: Decimal | None
    valor_cbs: Decimal | None
    ajuste_ibs: Decimal | None
    ajuste_cbs: Decimal | None
    valor_liquido: Decimal | None


def generate_nfse_report_xlsx(
    *,
    company: dict[str, object],
    tipo: str,
    data_inicial: str,
    data_final: str,
    situacao: str | None,
    query: str | None,
    documents: list[dict[str, object]],
) -> Path:
    rows = [_report_row(document) for document in documents]
    temp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_file.close()
    output_path = Path(temp_file.name)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Retencoes NFS-e"
    sheet.sheet_view.showGridLines = False

    headers = [
        "Número da NFS-e",
        "Situação",
        "Finalidade da NFS-e (finNFSe)",
        "Tipo de débito (tpNFSeDebito)",
        "Tipo de crédito (tpNFSeCredito)",
        "Valor total da nota",
        "IRRF retido (vRetIRRF)",
        "CSLL retida (vRetCSLL)",
        "ISSQN retido",
        "Contribuição previdenciária retida (vRetCP)",
        "Valor IBS (vIBSTot)",
        "Valor CBS (vCBS)",
        "Ajuste IBS (gIBSCBSAjuste/vIBS)",
        "Ajuste CBS (gIBSCBSAjuste/vCBS)",
        "Valor líquido (vLiq)",
    ]
    last_column = get_column_letter(len(headers))

    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = "RELATÓRIO DE RETENÇÕES DE NFS-e"
    sheet["A1"].font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="166534")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells(f"A2:{last_column}2")
    company_name = str(company.get("legal_name") or "Empresa")
    company_cnpj = _format_cnpj(str(company.get("cnpj") or ""))
    sheet["A2"] = f"{company_name} - CNPJ {company_cnpj}"

    sheet.merge_cells(f"A3:{last_column}3")
    type_label = "Emitidas" if tipo == "emitidas" else "Recebidas"
    sheet["A3"] = (
        f"Período: {_format_date(data_inicial)} a {_format_date(data_final)}"
        f" | Tipo: {type_label}"
    )

    sheet.merge_cells(f"A4:{last_column}4")
    filters = []
    if situacao and situacao != "todas":
        filters.append(f"Situação: {situacao.capitalize()}")
    if query:
        filters.append(f"Pesquisa: {query}")
    sheet["A4"] = "Filtros adicionais: " + (" | ".join(filters) if filters else "Nenhum")

    for cell_ref in ("A2", "A3", "A4"):
        sheet[cell_ref].font = Font(name="Arial", size=10, color="374151")
        sheet[cell_ref].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[2].height = 20
    sheet.row_dimensions[3].height = 20
    sheet.row_dimensions[4].height = 20

    header_row = 6
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=header)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="047857")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[header_row].height = 36

    first_data_row = header_row + 1
    for row_index, item in enumerate(rows, start=first_data_row):
        values = [
            item.numero,
            item.situacao,
            item.finalidade_nfse,
            item.tipo_debito,
            item.tipo_credito,
            _excel_number(item.valor_total),
            _excel_number(item.irrf_retido),
            _excel_number(item.csll_retida),
            _excel_number(item.issqn_retido),
            _excel_number(item.contribuicao_previdenciaria_retida),
            _excel_number(item.valor_ibs),
            _excel_number(item.valor_cbs),
            _excel_number(item.ajuste_ibs),
            _excel_number(item.ajuste_cbs),
            _excel_number(item.valor_liquido),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.font = Font(name="Arial", size=10, color="111827")
            cell.alignment = Alignment(
                horizontal="left" if column <= 5 else "right",
                vertical="center",
            )
            if column > 5:
                cell.number_format = MONEY_FORMAT
        if (row_index - first_data_row) % 2:
            for cell in sheet[row_index]:
                cell.fill = PatternFill("solid", fgColor="F0FDF4")
        sheet.row_dimensions[row_index].height = 20

    if not rows:
        sheet.cell(
            row=first_data_row,
            column=1,
            value="Nenhuma nota encontrada para os filtros informados.",
        )
        sheet.cell(row=first_data_row, column=1).font = Font(
            name="Arial",
            size=10,
            italic=True,
            color="6B7280",
        )

    last_data_row = max(first_data_row, header_row + len(rows))
    table = Table(displayName="RelatorioRetencoesNFSe", ref=f"A{header_row}:{last_column}{last_data_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    total_row = last_data_row + 2
    sheet.cell(row=total_row, column=1, value="TOTAIS")
    for column in range(6, len(headers) + 1):
        column_letter = get_column_letter(column)
        formula = f"=SUM({column_letter}{first_data_row}:{column_letter}{last_data_row})"
        sheet.cell(row=total_row, column=column, value=formula)
        sheet.cell(row=total_row, column=column).number_format = MONEY_FORMAT

    thin_green = Side(style="thin", color="86EFAC")
    for cell in sheet[total_row]:
        cell.font = Font(name="Arial", size=10, bold=True, color="14532D")
        cell.fill = PatternFill("solid", fgColor="DCFCE7")
        cell.border = Border(top=thin_green, bottom=thin_green)
        cell.alignment = Alignment(
            horizontal="left" if cell.column == 1 else "right",
            vertical="center",
        )
    sheet.row_dimensions[total_row].height = 22

    widths = [22, 16, 28, 28, 28, 20, 23, 23, 18, 43, 20, 20, 28, 28, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = f"A{first_data_row}"
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{last_data_row}"
    sheet.print_title_rows = f"1:{header_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "Página &P de &N"
    sheet.oddFooter.right.text = datetime.now().strftime("Gerado em %d/%m/%Y %H:%M")

    workbook.save(output_path)
    return output_path


def _report_row(document: dict[str, object]) -> ReportRow:
    xml_path = document.get("xml_path")
    if not isinstance(xml_path, str) or not Path(xml_path).exists():
        return ReportRow(
            numero=str(document.get("access_key") or "-"),
            situacao=_document_status(document),
            finalidade_nfse="",
            tipo_debito="",
            tipo_credito="",
            valor_total=_decimal_or_none(document.get("service_amount")),
            irrf_retido=None,
            csll_retida=None,
            issqn_retido=None,
            contribuicao_previdenciaria_retida=None,
            valor_ibs=None,
            valor_cbs=None,
            ajuste_ibs=None,
            ajuste_cbs=None,
            valor_liquido=_decimal_or_none(document.get("net_amount")),
        )

    try:
        root = ET.parse(xml_path).getroot()
    except (ET.ParseError, OSError):
        return ReportRow(
            numero=str(document.get("access_key") or "-"),
            situacao=_document_status(document),
            finalidade_nfse="",
            tipo_debito="",
            tipo_credito="",
            valor_total=_decimal_or_none(document.get("service_amount")),
            irrf_retido=None,
            csll_retida=None,
            issqn_retido=None,
            contribuicao_previdenciaria_retida=None,
            valor_ibs=None,
            valor_cbs=None,
            ajuste_ibs=None,
            ajuste_cbs=None,
            valor_liquido=_decimal_or_none(document.get("net_amount")),
        )

    inf_nfse = root.find("n:infNFSe", NAMESPACE)
    if inf_nfse is None and root.tag.endswith("infNFSe"):
        inf_nfse = root
    inf_dps = inf_nfse.find("n:DPS/n:infDPS", NAMESPACE) if inf_nfse is not None else None
    dps_values = inf_dps.find("n:valores", NAMESPACE) if inf_dps is not None else None
    nfse_values = inf_nfse.find("n:valores", NAMESPACE) if inf_nfse is not None else None
    ibscbs_dps = inf_dps.find("n:IBSCBS", NAMESPACE) if inf_dps is not None else None
    ibscbs_nfse = inf_nfse.find("n:IBSCBS", NAMESPACE) if inf_nfse is not None else None
    total_cibs = _find(ibscbs_nfse, "totCIBS")
    total_ibs = _find(total_cibs, "gIBS")
    total_cbs = _find(total_cibs, "gCBS")
    adjustment = _find(ibscbs_dps, "valores/trib/gIBSCBSAjuste")
    retention_type = _text(dps_values, "trib/tribMun/tpRetISSQN")
    issqn = _decimal(_text(nfse_values, "vISSQN")) if retention_type == "2" else Decimal("0")

    return ReportRow(
        numero=_text(inf_nfse, "nNFSe") or str(document.get("access_key") or "-"),
        situacao=_document_status(document),
        finalidade_nfse=_text(inf_dps, "finNFSe") or _text(ibscbs_dps, "finNFSe"),
        tipo_debito=_text(inf_dps, "tpNFSeDebito"),
        tipo_credito=_text(inf_dps, "tpNFSeCredito"),
        valor_total=_decimal(_text(dps_values, "vServPrest/vServ"))
        or _decimal_or_none(document.get("service_amount")),
        irrf_retido=_decimal(_text(dps_values, "trib/tribFed/vRetIRRF")),
        csll_retida=_decimal(_text(dps_values, "trib/tribFed/vRetCSLL")),
        issqn_retido=issqn,
        contribuicao_previdenciaria_retida=_decimal(
            _text(dps_values, "trib/tribFed/vRetCP")
        ),
        valor_ibs=_decimal(_text(total_ibs, "vIBSTot")),
        valor_cbs=_decimal(_text(total_cbs, "vCBS")),
        ajuste_ibs=_decimal(_text(adjustment, "vIBS")),
        ajuste_cbs=_decimal(_text(adjustment, "vCBS")),
        valor_liquido=_decimal(_text(nfse_values, "vLiq"))
        or _decimal_or_none(document.get("net_amount")),
    )


def _document_status(document: dict[str, object]) -> str:
    return str(document.get("status") or "Autorizada")


def _text(node: ET.Element | None, path: str) -> str:
    if node is None:
        return ""
    found = _find(node, path)
    return (found.text or "").strip() if found is not None else ""


def _find(node: ET.Element | None, path: str) -> ET.Element | None:
    if node is None:
        return None
    return node.find(f"n:{path.replace('/', '/n:')}", NAMESPACE)


def _decimal(value: str) -> Decimal:
    try:
        return Decimal(value) if value else Decimal("0")
    except InvalidOperation:
        return Decimal("0")


def _decimal_or_none(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return None


def _excel_number(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def _format_date(value: str) -> str:
    return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")


def _format_cnpj(value: str) -> str:
    digits = "".join(character for character in value if character.isdigit())
    if len(digits) != 14:
        return value or "-"
    return (
        f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/"
        f"{digits[8:12]}-{digits[12:]}"
    )
